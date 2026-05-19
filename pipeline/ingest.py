from datetime import datetime, timezone
from pathlib import Path

from core.models import InvoiceRecord, NormalizedInput

_EXTENSION_MAP = {
    ".txt": "plain_text",
    ".pdf": "pdf",
    ".html": "html",
    ".htm": "html",
    ".xlsx": "excel",
    ".xls": "excel",
    ".csv": "excel",
    ".png": "scan",
    ".jpg": "scan",
    ".jpeg": "scan",
    ".tiff": "scan",
    ".tif": "scan",
}


def normalize_to_text(source: Path, source_type: str | None = None) -> NormalizedInput:
    """
    Converteert een brondocument naar genormaliseerde tekst voor de extractie-pipeline.

    Ondersteund: plain_text, pdf, html, excel
    Niet geïmplementeerd: scan — gebruik pytesseract of externe OCR-API.
    """
    resolved_type = source_type or _EXTENSION_MAP.get(source.suffix.lower(), "plain_text")

    if resolved_type == "plain_text":
        return _read_plain_text(source)
    elif resolved_type == "pdf":
        return _read_pdf(source)
    elif resolved_type == "html":
        return _read_html(source)
    elif resolved_type == "excel":
        return _read_excel(source)
    elif resolved_type == "scan":
        raise NotImplementedError(
            "Scan/afbeelding-ingest is nog niet geïmplementeerd.\n"
            "Aanbevolen aanpak: gebruik `pytesseract` met Tesseract-OCR voor lokale verwerking, "
            "of een externe OCR-API (bijv. Google Cloud Vision, Azure Form Recognizer). "
            "Sla `ocr_confidence` op in metadata."
        )
    else:
        raise NotImplementedError(
            f"Bestandstype '{resolved_type}' wordt niet herkend. "
            f"Gebruik --source-type om het type expliciet op te geven."
        )


def _read_plain_text(source: Path) -> NormalizedInput:
    text = source.read_text(encoding="utf-8")
    return NormalizedInput(
        text=text,
        source_file=str(source),
        source_type="plain_text",
        metadata={"encoding": "utf-8", "size_bytes": source.stat().st_size},
    )


def _read_pdf(source: Path) -> NormalizedInput:
    import pdfplumber
    pages = []
    with pdfplumber.open(source) as pdf:
        page_count = len(pdf.pages)
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
    return NormalizedInput(
        text="\n\n".join(pages),
        source_file=str(source),
        source_type="pdf",
        metadata={"page_count": page_count, "pages_with_text": len(pages)},
    )


def _read_html(source: Path) -> NormalizedInput:
    from bs4 import BeautifulSoup
    html = source.read_text(encoding="utf-8")
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "nav", "header", "footer"]):
        tag.decompose()
    title = soup.title.string.strip() if soup.title else ""
    text = soup.get_text(separator="\n", strip=True)
    return NormalizedInput(
        text=text,
        source_file=str(source),
        source_type="html",
        metadata={"title": title},
    )


def _read_excel(source: Path) -> NormalizedInput:
    import openpyxl
    wb = openpyxl.load_workbook(source, data_only=True)
    sheet_names = wb.sheetnames
    parts = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
    return NormalizedInput(
        text="\n\n".join(parts),
        source_file=str(source),
        source_type="excel",
        metadata={"sheet_names": sheet_names},
    )


def ingest(file_path: str | Path, source_type: str | None = None) -> InvoiceRecord:
    path = Path(file_path)
    normalized = normalize_to_text(path, source_type=source_type)
    now = datetime.now(timezone.utc)
    run_id = now.strftime("%Y%m%d-%H%M%S") + f"-{path.stem}"
    return InvoiceRecord(
        run_id=run_id,
        source_file=str(path),
        source_type=normalized.source_type,
        raw_text=normalized.text,
        tenant_slug=None,
        status="new",
        created_at=now.isoformat(),
    )
